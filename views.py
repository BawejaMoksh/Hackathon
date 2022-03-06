from django.shortcuts import render
import openpyxl


def index(request):
    if "GET" == request.method:
        return render(request, 'myapp/web.html', {})
    # else:
    #     excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size
    if "POST" == request.method:
        wb = openpyxl.load_workbook('TestdataNew.xlsx')

        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Further sheet"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)

    return render(request, 'myapp/web.html', {"excel_data": excel_data})
